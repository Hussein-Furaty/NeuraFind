from pathlib import Path

from openpyxl import load_workbook


class XlsxParser:
    """Extracts text content from XLSX spreadsheets."""

    def extract_text(self, file_path: str | Path) -> str:
        path = Path(file_path)

        if not path.exists():
            raise FileNotFoundError(f"Spreadsheet does not exist: {path}")

        if path.suffix.lower() != ".xlsx":
            raise ValueError(f"Unsupported file type: {path.suffix}")

        workbook = load_workbook(path, data_only=True)

        text_parts: list[str] = []

        for sheet in workbook.worksheets:
            text_parts.append(f"Sheet: {sheet.title}")

            for row in sheet.iter_rows(values_only=True):
                values = [str(cell) for cell in row if cell is not None]

                if values:
                    text_parts.append(" | ".join(values))

        return "\n".join(text_parts)